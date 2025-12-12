import pandas as pd
import openpyxl
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

tmplt=r'D:\WorkDir\ProjSpace\M2114_Wisting\Sima\_Res_SD_'

# List of sheet names
DesignBookB = [
    'B_ULS_Sd', 'B_ALS1F_Sd100yr', 'B_ALS2F_Sd10yr',
    'B_ULS_Sd10kyr', 'B_ALS1F_Sd1yr', 'B_ALS2F_Sd1yr', 'B_ALS3F_Sd1yr',
]
DesignBookL = [
    'L_ULS_Sd', 'L_ALS1F_Sd100yr', 'L_ALS2F_Sd10yr',
    'L_ULS_Sd10kyr', 'L_ALS1F_Sd1yr', 'L_ALS2F_Sd1yr', 'L_ALS3F_Sd1yr'
]

# Load source Excel file
#source_file = r'D:\WorkDir\ProjSpace\M2114_Wisting\Metocean\DesignBook_LowMLBE.xlsx'  # Replace with your actual source file
source_file = r'D:\WorkDir\ProjSpace\M2114_Wisting\Metocean\DesignBook_Cstudy.xlsx'  # Replace with your actual source file
wb_source = pd.ExcelFile(source_file)
'''
------------------------------------------------------
 Method	             |   Description
 wb.active	         |   Active sheet when last saved
 wb['Sheet1']	     |   Access by exact name
 wb.sheetnames	     |   List of sheet names
 wb.worksheets	     |   List of worksheet objects
 wb.create_sheet()	 |   Add a new sheet
 del wb['Sheet1']    |   Delete a sheet
'''
# Transformation function
def transform_code(code):
    if not isinstance(code, str):
        return code
    # Remove the last underscore and the two digits
    code = re.sub(r'_(\d{2})$', '', code)
    # Split by underscore
    parts = code.split('_')
    if len(parts) > 1:
        # Leave the first part intact, split the second part into 2-digit segments
        first = parts[0]
        rest = parts[1]
        rest_split = [rest[i:i+2] for i in range(0, len(rest), 2)]
        return '_'.join([first] + rest_split)
    return code

def write_design_value_summary(design_space:str, design_sheets:list=DesignBookB, design_book:str=source_file,start_row:int = 93):
    # Columns and corresponding Excel labels
    columns = ['T', 'AG', 'AH', 'AI', 'AJ', 'AK']
    start_ro = 5
    end_ro = 16
    
    # Prepare results list
    summary_data = []
    
    for file in design_sheets:
        # Load workbook and sheet
        wb = load_workbook(filename= design_space + file +'.xlsx', data_only=True)
        sheet = wb.active
        
        maxima = []
        for col in columns:
            col_values = []
            for row_num in range(start_ro, end_ro + 1):
                cell_value = sheet[f"{col}{row_num}"].value
                if cell_value is not None:
                    col_values.append(cell_value)
            max_val = max(col_values) if col_values else None
            maxima.append(max_val)
    
        
        summary_data.append([file.replace('.xlsx', '')] + maxima)
    
    # Create DataFrame for summary
    summary_df = pd.DataFrame(summary_data, columns=['Ballast']+columns)
    
    # Open existing workbook
    wb_target = load_workbook(filename=design_book)
    ws_target = wb_target["Loadcases_Sd"]
    
    # Write the DataFrame to A93
    #start_row = 93
    start_col = 1
    
    for i, row in enumerate(dataframe_to_rows(summary_df, index=False, header=False), start=start_row):
        for j, value in enumerate(row, start=start_col):
            ws_target.cell(row=i, column=j).value = value  # only update value, not style Writing into pre-formatted templates
            #ws_target.cell(row=i, column=j, value=value) # Writing to new/empty sheets
    # Save the result
    wb_target.save(design_book)
    # Show table to user
    #import ace_tools as tools; tools.display_dataframe_to_user(name="DesignBook Maxima Summary", dataframe=summary_df)

def design_resxls_gen(template:str=tmplt, design_sheets:list=DesignBookB):
# Process each sheet
    # Open target workbook
    tmplt_file = f'{template}{design_sheets[0][0]}.xlsx'
    #target_file = f'{sheet_name}.xlsx'
#    wb_target = openpyxl.load_workbook(tmplt_file)
#    ws = wb_target.active
    for sheet_name in design_sheets:
        # Reload a fresh copy of the template for each iteration
        wb_target = openpyxl.load_workbook(tmplt_file)
        ws = wb_target.active

        # Read the range Q2:Q13 from the source sheet
        df = wb_source.parse(sheet_name, usecols='Q', skiprows=1, nrows=12, header=None)
        print(df.head(12))
        df_transformed = df.iloc[:, 0].apply(transform_code)
   
        # Write to range S5:S16
        for i, value in enumerate(df_transformed, start=5):
            ws[f'S{i}'] = value
    
        # Set formula in AA4
        ws['AA3'] = '=PreTen' + sheet_name[0]
        ws['AE3'] = '=SUMPRODUCT(--(AA5:AA244 = AE5:AE244)) = ROWS(AA5:AA244)'
         # Replace "B5" with "PreTenB" in formulas in AA5:AA244 and AE5:AE244
        for row in range(5, 10):
            for col in ['AA', 'AE']:
                cell = ws[f'{col}{row}']
                if isinstance(cell.value, str) and f'B{row}' in cell.value:
                    # Replace only B{row} — e.g., B5 in row 5 — with PreTenL
                    updated_formula = cell.value.replace(f'B{row}', 'PreTen' + sheet_name[0])
                    if cell.value != updated_formula:
                        cell.value = updated_formula
            # Save the updated workbook
        wb_target.save(f'{sheet_name}.xlsx')
    
    print("Data transfer and transformation complete.")

def sre_lc_file_gen(design_book:str=source_file, design_sheets:list=DesignBookB,min_row:int=2, max_row:int=85, min_col:int=31, max_col:int=35):
    import os
    # Columns and corresponding Excel labels
    #columns = ['AE', 'AF', 'AG', 'AH', 'AI']
    #start_ro = 2
    #end_ro = 85
    
    # Prepare results list
    
    for file in design_sheets:
        # Load workbook and sheet
        wb = load_workbook(filename=design_book, data_only=True)
        sheet = wb[file]
        
        # Extract AE2:AI85 and save as text
        text_output_path = os.path.dirname(source_file) + '\\'+ file + '.txt'
        with open(text_output_path, 'w') as f:
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                values = [str(cell.value) if cell.value is not None else "" for cell in row]
                f.write('\t'.join(values) + '\n')

        # (Optional) Continue with value summary logic here...

    print("Text export complete for all files.")
    
def offset_minmax(base_path:str = r"D:\WorkDir\ProjSpace\M2114_Wisting\Sima\SRE_LowMLBE_Loaded\L_ULS_Screen"):
    """
    Collect the offset data from "key_sima_noddis.txt.xlsx" and write into a summary file.
    """
    import os
    import pandas as pd
    from tqdm import tqdm

    # 基础路径配置

    output_path = os.path.join(base_path, "extreme_values_summary.xlsx")

    # 关键参数定义
    start_row = 2007     # 跳过前1000秒后的起始行（时间步长0.5秒）
    interval_size = 10080 * 2  # 每个时间区间的大小（10080 * 2=20160行）

    # 列索引配置（B列=1, C列=2）
    cols_to_read = [1, 2]
    col_names = ['B', 'C']

    # 准备结果存储
    results = []

    # 遍历144个文件夹（rLC1 到 rLC144）
    for i in tqdm(range(1, 145)):
        folder_name = f"rLC{i}"
        file_path = os.path.join(base_path, folder_name, "node_1", "key_sima_noddis.txt.xlsx")
    
        try:
            # 读取Excel文件（跳过前6行标题，只读取B列和C列）
            df = pd.read_excel(
                file_path,
                sheet_name='TimeSeries',
                skiprows=6,
                usecols=cols_to_read,
                header=None
            )
            df.columns = col_names
        
            # 转换为数值类型（处理可能的格式问题）
            df = df.apply(pd.to_numeric, errors='coerce')
        
            # 计算三个时间区间的行范围
            intervals = [
                (start_row, start_row + interval_size),
                (start_row + interval_size + 1, start_row + 2 * interval_size),
                (start_row + 2 * interval_size + 1, start_row + 3 * interval_size)
            ]
        
            # 检查数据是否足够
            if df.shape[0] < intervals[2][1]:
                print(f"⚠️ 数据不足: {folder_name} (需要{intervals[2][1]}行, 实际{df.shape[0]}行)")
                continue
        
            # 计算每个区间的极值
            folder_results = {'Folder': folder_name}
            for idx, (start, end) in enumerate(intervals, 1):
                interval_data = df.iloc[start-1:end-1]  # 转换为0-based索引
            
                for col in col_names:
                    col_data = interval_data[col].dropna()
                    if not col_data.empty:
                        folder_results[f'Interval{idx}_{col}_min'] = col_data.min()
                        folder_results[f'Interval{idx}_{col}_max'] = col_data.max()
                    else:
                        folder_results[f'Interval{idx}_{col}_min'] = None
                        folder_results[f'Interval{idx}_{col}_max'] = None
        
            results.append(folder_results)
    
        except Exception as e:
            print(f"❌ 处理失败 {folder_name}: {str(e)}")
            results.append({'Folder': folder_name, 'Error': str(e)})

    # 保存结果到Excel
    if results:
        result_df = pd.DataFrame(results)
        result_df.to_excel(output_path, index=False)
        print(f"\n✅ 处理完成! 结果已保存至: {output_path}")
        print(f"成功处理: {len([r for r in results if 'Error' not in r])}/{len(results)}个文件夹")
    else:
        print("❌ 未找到有效数据")        


#=XLOOKUP(1;(INDEX(SFactors;;2)=10^LEFT(A5;1))*(INDEX(SFactors;;3)=(LEN(A5)-LEN(SUBSTITUTE(A5;"_";""))));INDEX(SFactors;;4))*PreTenB+(C5-PreTenB)*XLOOKUP(1;(INDEX(SFactors;;2)=10^LEFT(A5;1))*(INDEX(SFactors;;3)=(LEN(A5)-LEN(SUBSTITUTE(A5;"_";""))));INDEX(SFactors;;5))	=AA5/MBL	"=XLOOKUP(1;
#    (INDEX(SFactors;;2)=10^LEFT(A5;1)) *
#    (INDEX(SFactors;;3)=(LEN(A5)-LEN(SUBSTITUTE(A5;""_"";""""))));
#    INDEX(SFactors;;4)
#)"	=XLOOKUP(1;(INDEX(SFactors;;2)=10^LEFT(A5;1))*(INDEX(SFactors;;3)=(LEN(A5)-LEN(SUBSTITUTE(A5;"_";""))));INDEX(SFactors;;5))	=PreTenB*AC5+(C5-PreTenB)*AD5

if __name__ == "__main__":
    #sre_lc_file_gen(design_sheets=DesignBookB,design_book=source_file)
    #sre_lc_file_gen(design_sheets=DesignBookL,design_book=source_file)
    design_resxls_gen(tmplt,DesignBookB)
    design_resxls_gen(tmplt,DesignBookL)
    #write_design_value_summary(design_space='D:\\WorkDir\\ProjSpace\\M2114_Wisting\\Sima\\SRE_LowMLBE_Ballast\\', design_sheets=DesignBookB,design_book=source_file,start_row=93)
    #write_design_value_summary(design_space='D:\\WorkDir\\ProjSpace\\M2114_Wisting\\Sima\\SRE_LowMLBE_Loaded\\', design_sheets=DesignBookL,design_book=source_file,start_row=93+len(DesignBookB)+1)
    #offset_minmax()